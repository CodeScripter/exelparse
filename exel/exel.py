import openpyxl
import xlrd
def exelparse(filename):
	if filename == "11Tax Savings Calculator - TCs.xlsx":
		path = filename
		wb_obj = openpyxl.load_workbook(path)
		# book = xlrd.open_workbook(path)
		# sheet = book.sheet_by_name('Sanity')

		sheet_obj = wb_obj.active
		cell_obj= sheet_obj.cell(row = 2, column = 2)
		Annual_Income = cell_obj.value
		cell_obj = sheet_obj.cell(row = 5, column = 2)
		Monthly_Rent = cell_obj.value
		cell_obj = sheet_obj.cell(row = 6, column = 2)
		Monthly_HRA = cell_obj.value
		cell_obj= sheet_obj.cell(row = 7, column = 2)
		Tution_Fees = cell_obj.value
		cell_obj = sheet_obj.cell(row = 8, column = 2)
		Interest_Paid_on_Education_Loan = cell_obj.value
		cell_obj = sheet_obj.cell(row = 14, column = 2)
		Monthly_EPF = cell_obj.value
		cell_obj= sheet_obj.cell(row = 15, column = 2)
		L_I_C = cell_obj.value
		cell_obj = sheet_obj.cell(row = 16, column = 2)
		P_P_F = cell_obj.value
		cell_obj = sheet_obj.cell(row = 17, column = 2)
		Tax_savings_mutual_funds = cell_obj.value
		cell_obj= sheet_obj.cell(row = 18, column = 2)
		Tax_savings_FD = cell_obj.value
		cell_obj = sheet_obj.cell(row = 19, column = 2)
		N_P_S = cell_obj.value
		cell_obj = sheet_obj.cell(row = 20, column = 2)
		Health_Insurance_Premium_for_Self = cell_obj.value
		cell_obj= sheet_obj.cell(row = 21, column = 2)
		Health_Insurance_Premium_for_Parents = cell_obj.value
		cell_obj = sheet_obj.cell(row = 23, column = 2)
		Tax_old_regime = cell_obj.value
		cell_obj = sheet_obj.cell(row = 24, column = 2)
		Tax_new_regime = cell_obj.value
		print(Annual_Income, "\n", Monthly_Rent, "\n", Monthly_HRA, "\n", Tution_Fees, "\n", Interest_Paid_on_Education_Loan,
		"\n", Monthly_EPF, "\n", L_I_C, "\n", P_P_F, "\n", Tax_savings_mutual_funds, "\n", Tax_savings_FD, "\n", N_P_S, "\n", 
		Health_Insurance_Premium_for_Self, "\n", Health_Insurance_Premium_for_Parents)

	elif filename ==  "Retirment Calculator final.xlsx":
		path = filename
		wb_obj = openpyxl.load_workbook(path)
		sheet_obj = wb_obj.active
		cell_obj_target_amt = sheet_obj.cell(row = 8, column = 2)
		Target_Amount = cell_obj_target_amt.value
		cell_obj_target_time = sheet_obj.cell(row = 10, column = 2)
		Target_Time = cell_obj_target_time.value
		Target_Amount = int(Target_Amount)
		Target_Time = int(Target_Time/12)
		print(Target_Amount,Target_Time)

	elif filename == "Tax Savings Calculator - TCs.xlsx":
		path = filename
		wb_obj = openpyxl.load_workbook(path)
		print(wb_obj.get_sheet_names())  
		# sheet = wb_obj.get_sheet_by_name("Sanity")  
		# sheet_obj = sheet.active
		# cell_obj= sheet_obj.cell(row = 8, column = 2)
		# Target_Amount = cell_obj.value
		# cell_obj = sheet_obj.cell(row = 10, column = 2)
		# Target_Time = cell_obj.value
		# Target_Amount = int(Target_Amount)
		# Target_Time = int(Target_Time/12)
		# print(Target_Amount,Target_Time)

	return Annual_Income, Monthly_Rent, Monthly_HRA, Tution_Fees, Interest_Paid_on_Education_Loan, Monthly_EPF, L_I_C, P_P_F, Tax_savings_mutual_funds, Tax_savings_FD, N_P_S, Health_Insurance_Premium_for_Self, Health_Insurance_Premium_for_Parents
exelparse("Tax Savings Calculator - TCs.xlsx")
